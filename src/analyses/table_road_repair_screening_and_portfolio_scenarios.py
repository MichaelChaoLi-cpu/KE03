"""Generate the combined road-repair screening and portfolio table.

Plan: integrate the two positive single-section candidates with the four
count-constrained, jointly rerouted portfolios in one compact main-text table.

Framework: AnaSOP Section 6.7 and workflow step 10. Single-section rows are
marginal results with every other primary closure retained. Portfolio rows are
joint-rerouting results for K in {1, 2, 3, 5}; they are not sums of candidate
benefits, engineering repair orders, or cost-optimal plans.
"""

from __future__ import annotations

from pathlib import Path
import shutil
import subprocess
import tempfile

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image, ImageChops


ROOT = Path(__file__).resolve().parents[2]
CANDIDATES = (
    ROOT
    / "data"
    / "processed"
    / "decision"
    / "road_repair_candidate_benefits_preprocessed.parquet"
)
PORTFOLIO_SUMMARY = (
    ROOT
    / "data"
    / "processed"
    / "decision"
    / "road_repair_portfolio_summary_preprocessed.parquet"
)
PORTFOLIO_SCENARIOS = (
    ROOT
    / "data"
    / "processed"
    / "decision"
    / "road_repair_portfolio_scenarios_preprocessed.parquet"
)
OUTPUT = (
    ROOT
    / "data"
    / "results"
    / "tables"
    / "Table_road_repair_screening_and_portfolio_scenarios.xlsx"
)
PREVIEW = (
    ROOT
    / "data"
    / "exp"
    / "table_previews"
    / "Table_road_repair_screening_and_portfolio_scenarios.png"
)

TITLE = "Road-Repair Screening and Portfolio Scenarios"
HEADERS = [
    "Analysis level",
    "Selected road section(s)",
    "Restored\nsettlements",
    "Population\nreconnected",
    "Finite improvement\n(person-minutes)",
    "Marginal population\nreconnected",
    "Marginal finite improvement\n(person-minutes)",
    "Structural-scenario\nretention",
]


def candidate_ids(value: object) -> list[str]:
    return [item for item in str(value).split(";") if item]


def compact_section(section: str) -> str:
    replacements = {
        "Pasang Lhamu Highway (OSM 379104232)": "Pasang Lhamu Highway (379104232)",
        "Unnamed unclassified road (OSM 533216634)": "Unnamed road (533216634)",
        "Unnamed tertiary road (OSM 1463147995)": "Unnamed tertiary (1463147995)",
        "Unnamed residential road (OSM 121203554)": "Unnamed residential (121203554)",
        "Trishuli Bridge (OSM 1219434229)": "Trishuli Bridge (1219434229)",
    }
    return replacements.get(section.strip(), section.strip().replace("OSM ", ""))


def compact_section_set(value: object) -> str:
    return "\n".join(compact_section(item) for item in str(value).split(";") if item)


def load_and_validate() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Validate the candidate surface, primary portfolios, and robustness grid."""
    candidates = pd.read_parquet(CANDIDATES)
    summary = pd.read_parquet(PORTFOLIO_SUMMARY).sort_values(
        "Repair Portfolio Size (sections)"
    )
    scenarios = pd.read_parquet(PORTFOLIO_SCENARIOS)
    positive = candidates.loc[
        candidates["Is Critical Road Section"].fillna(False)
    ].sort_values("Primary Repair Benefit Rank")

    if len(candidates) != 194 or len(positive) != 2:
        raise RuntimeError("Expected 194 screened candidates and two positive candidates")
    if positive["Primary Repair Benefit Rank"].tolist() != [1, 2]:
        raise RuntimeError("Positive candidate order does not match the screening tuple")
    if summary["Repair Portfolio Size (sections)"].tolist() != [1, 2, 3, 5]:
        raise RuntimeError("Expected the four pre-specified portfolio sizes")
    if len(scenarios) != 768:
        raise RuntimeError("Expected 768 portfolio-scenario records")
    if not scenarios.groupby("Repair Portfolio Size (sections)").size().eq(192).all():
        raise RuntimeError("Each portfolio must be evaluated in 192 scenarios")

    primary = scenarios.loc[scenarios["Primary Scenario"].fillna(False)].copy()
    primary = primary.sort_values("Repair Portfolio Size (sections)")
    if len(primary) != 4:
        raise RuntimeError("Expected one primary result for each portfolio size")
    for _, summary_row in summary.iterrows():
        k = int(summary_row["Repair Portfolio Size (sections)"])
        primary_row = primary.loc[
            primary["Repair Portfolio Size (sections)"].eq(k)
        ].iloc[0]
        checks = [
            (
                primary_row["Portfolio Population Reconnected"],
                summary_row["Primary Portfolio Population Reconnected"],
            ),
            (
                primary_row[
                    "Portfolio-Weighted Finite Travel-Time Improvement (person-minutes)"
                ],
                summary_row[
                    "Primary Portfolio-Weighted Finite Travel-Time Improvement (person-minutes)"
                ],
            ),
            (
                primary_row["Portfolio Structural-Scenario Retention"],
                summary_row["Portfolio Structural-Scenario Retention"],
            ),
        ]
        if any(not np.isclose(left, right) for left, right in checks):
            raise RuntimeError(f"Primary and summary portfolio values disagree for K={k}")

    portfolio_sets = [set(candidate_ids(value)) for value in summary["Selected Road Repair Candidate IDs"]]
    if any(not portfolio_sets[i - 1].issubset(portfolio_sets[i]) for i in range(1, 4)):
        raise RuntimeError("Portfolio candidate sets are not nested")
    positive_ids = positive["Road Repair Candidate ID"].astype(str).tolist()
    if portfolio_sets[0] != {positive_ids[0]} or portfolio_sets[1] != set(positive_ids):
        raise RuntimeError("K=1 and K=2 do not align with the two positive candidates")
    if not summary["Portfolio Structural-Scenario Retention"].apply(
        lambda value: np.isclose(value, 2 / 3)
    ).all():
        raise RuntimeError("Unexpected structural-scenario retention")
    return positive, summary, primary


def build_rows(
    positive: pd.DataFrame,
    summary: pd.DataFrame,
    primary: pd.DataFrame,
) -> list[list[object]]:
    rows: list[list[object]] = []

    for order, (_, row) in enumerate(positive.iterrows(), start=1):
        population = int(round(float(row["Population Reconnected"])))
        finite = int(
            round(
                float(
                    row[
                        "Population-Weighted Finite Travel-Time Improvement (person-minutes)"
                    ]
                )
            )
        )
        rows.append(
            [
                f"Single section {order}",
                compact_section(str(row["Critical Road Section"])),
                int(row["Settlements Reconnected"]),
                population,
                finite,
                population,
                finite,
                "n/a",
            ]
        )

    preceding_population = 0.0
    preceding_finite = 0.0
    for _, row in summary.iterrows():
        k = int(row["Repair Portfolio Size (sections)"])
        primary_row = primary.loc[
            primary["Repair Portfolio Size (sections)"].eq(k)
        ].iloc[0]
        population = float(row["Primary Portfolio Population Reconnected"])
        finite = float(
            row[
                "Primary Portfolio-Weighted Finite Travel-Time Improvement (person-minutes)"
            ]
        )
        rows.append(
            [
                f"Portfolio K={k}",
                compact_section_set(row["Selected Critical Road Sections"]),
                int(primary_row["Settlements Reconnected by Portfolio"]),
                int(round(population)),
                int(round(finite)),
                int(round(population - preceding_population)),
                int(round(finite - preceding_finite)),
                float(row["Portfolio Structural-Scenario Retention"]),
            ]
        )
        preceding_population = population
        preceding_finite = finite

    if len(rows) != 6 or any(len(row) != 8 for row in rows):
        raise RuntimeError("AnaSOP requires exactly 6 rows and 8 columns")
    if rows[4][5:7] != [0, 0] or rows[5][5:7] != [0, 0]:
        raise RuntimeError("The K=3 and K=5 benefit plateau is not preserved")
    return rows


def build_workbook(rows: list[list[object]]) -> None:
    """Create the authoritative XLSX in the approved compact table style."""
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    PREVIEW.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Repair Scenarios"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A3"

    sheet.merge_cells("A1:H1")
    sheet["A1"] = TITLE
    sheet["A1"].font = Font(name="Aptos Display", size=17, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="17324D")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 32

    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=2, column=column, value=header)
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3182BD")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 48

    body_font = Font(name="Aptos", size=9.5, color="25313A")
    inner = Side(style="thin", color="D6DEE5")
    outer = Side(style="medium", color="535D66")
    for row_idx, values in enumerate(rows, start=3):
        portfolio_row = row_idx >= 5
        base_fill = "EEF4FA" if portfolio_row else "F4F7F9"
        if row_idx % 2 == 0:
            base_fill = "FFFFFF"
        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.fill = PatternFill("solid", fgColor=base_fill)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = Border(bottom=inner)
        sheet.cell(row=row_idx, column=1).font = Font(
            name="Aptos", size=9.5, bold=True, color="17324D"
        )
        sheet.cell(row=row_idx, column=1).fill = PatternFill(
            "solid", fgColor="D9EAD3" if not portfolio_row else "DDEBF7"
        )
        for col_idx in (3, 4, 5, 6, 7, 8):
            sheet.cell(row=row_idx, column=col_idx).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        for col_idx in (3, 4, 5, 6, 7):
            sheet.cell(row=row_idx, column=col_idx).number_format = "#,##0"
        if isinstance(values[7], float):
            sheet.cell(row=row_idx, column=8).number_format = "0%"

        if portfolio_row and int(values[5]) == 0 and int(values[6]) == 0:
            for col_idx in (6, 7):
                sheet.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    "solid", fgColor="FCE4D6"
                )
        if portfolio_row:
            sheet.cell(row=row_idx, column=8).fill = PatternFill(
                "solid", fgColor="FFF2CC"
            )
        section_count = str(values[1]).count("\n") + 1
        sheet.row_dimensions[row_idx].height = max(40, 18 + section_count * 16)

    widths = [20, 43, 17, 20, 24, 23, 27, 21]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    table = Table(displayName="RepairScreeningPortfolios", ref="A2:H8")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    for col_idx in range(1, 9):
        sheet.cell(row=2, column=col_idx).border = Border(top=outer, bottom=outer)
        sheet.cell(row=5, column=col_idx).border = Border(
            top=outer,
            bottom=sheet.cell(row=5, column=col_idx).border.bottom,
            left=sheet.cell(row=5, column=col_idx).border.left,
            right=sheet.cell(row=5, column=col_idx).border.right,
        )
        sheet.cell(row=8, column=col_idx).border = Border(bottom=outer)
    for row_idx in range(2, 9):
        left = sheet.cell(row=row_idx, column=1)
        right = sheet.cell(row=row_idx, column=8)
        left.border = Border(
            left=outer,
            top=left.border.top,
            bottom=left.border.bottom,
        )
        right.border = Border(
            right=outer,
            top=right.border.top,
            bottom=right.border.bottom,
        )

    sheet.auto_filter.ref = "A2:H8"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.print_area = "A1:H8"
    sheet.page_margins.left = 0.2
    sheet.page_margins.right = 0.2
    sheet.page_margins.top = 0.25
    sheet.page_margins.bottom = 0.25
    sheet.page_margins.header = 0
    sheet.page_margins.footer = 0
    workbook.save(OUTPUT)


def render_preview() -> None:
    """Render the workbook print area to a tightly cropped PNG."""
    soffice = shutil.which("soffice") or "/opt/homebrew/bin/soffice"
    pdftoppm = shutil.which("pdftoppm")
    if not Path(soffice).exists() or pdftoppm is None:
        raise RuntimeError("LibreOffice and pdftoppm are required for PNG rendering")

    with tempfile.TemporaryDirectory(prefix="ke03-table-preview-") as temp_name:
        temp_dir = Path(temp_name)
        profile = temp_dir / "lo-profile"
        subprocess.run(
            [
                soffice,
                f"-env:UserInstallation={profile.as_uri()}",
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                str(temp_dir),
                str(OUTPUT),
            ],
            check=True,
            capture_output=True,
            text=True,
        )
        pdf = temp_dir / f"{OUTPUT.stem}.pdf"
        preview_stem = temp_dir / "preview"
        subprocess.run(
            ["pdftoppm", "-png", "-singlefile", "-r", "180", str(pdf), str(preview_stem)],
            check=True,
            capture_output=True,
            text=True,
        )
        rendered = Image.open(preview_stem.with_suffix(".png")).convert("RGB")
        background = Image.new("RGB", rendered.size, "white")
        content_bounds = ImageChops.difference(rendered, background).getbbox()
        if content_bounds is None:
            raise RuntimeError("Rendered PNG contains no visible table content")
        left, top, right, bottom = content_bounds
        padding = 24
        rendered.crop(
            (
                max(0, left - padding),
                max(0, top - padding),
                min(rendered.width, right + padding),
                min(rendered.height, bottom + padding),
            )
        ).save(PREVIEW, dpi=(180, 180))


def validate_outputs(rows: list[list[object]]) -> None:
    """Validate workbook values, dimensions, errors, and PNG output."""
    workbook = load_workbook(OUTPUT, data_only=False, read_only=False)
    sheet = workbook["Repair Scenarios"]
    observed_headers = [sheet.cell(row=2, column=col).value for col in range(1, 9)]
    observed_rows = [
        [sheet.cell(row=row, column=col).value for col in range(1, 9)]
        for row in range(3, 9)
    ]
    if observed_headers != HEADERS:
        raise RuntimeError("Workbook headers do not match the combined table plan")
    for expected, observed in zip(rows, observed_rows, strict=True):
        for expected_value, observed_value in zip(expected, observed, strict=True):
            if isinstance(expected_value, float):
                if not np.isclose(expected_value, observed_value):
                    raise RuntimeError("Workbook numeric values changed during serialization")
            elif expected_value != observed_value:
                raise RuntimeError("Workbook values do not match the combined table plan")
    for row in sheet.iter_rows(min_row=1, max_row=8, min_col=1, max_col=8):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("#"):
                raise RuntimeError(f"Spreadsheet error value in {cell.coordinate}: {cell.value}")
            if cell.data_type == "f":
                raise RuntimeError(f"Unexpected formula in {cell.coordinate}")
    if not PREVIEW.exists() or PREVIEW.stat().st_size < 10_000:
        raise RuntimeError("PNG preview is missing or unexpectedly small")


def main() -> None:
    positive, summary, primary = load_and_validate()
    rows = build_rows(positive, summary, primary)
    build_workbook(rows)
    render_preview()
    validate_outputs(rows)
    print(f"Wrote {OUTPUT.relative_to(ROOT)}")
    print(f"Wrote {PREVIEW.relative_to(ROOT)}")
    print("Validated 6 rows x 8 columns; 2 candidates, 4 portfolios, 768 scenario records")


if __name__ == "__main__":
    main()
