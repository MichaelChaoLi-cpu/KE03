"""Generate the Evidence Base and Validation Status table and PNG preview.

The table implements the evidence-audit table in AnaSOP Section 8. It audits
eight evidence or validation domains against the role for which each layer is
used in the applied disaster analysis. Validation language is deliberately
role-specific: it does not promote remote-sensing interpretations to field
observations, population exposure to casualties, or repair screening to an
engineering recommendation.
"""

from __future__ import annotations

from pathlib import Path
import re
import shutil
import subprocess
import tempfile

import geopandas as gpd
import numpy as np
import pandas as pd
import rasterio
from rasterio.warp import Resampling, reproject
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image, ImageChops


ROOT = Path(__file__).resolve().parents[2]
OUTPUT = (
    ROOT
    / "data"
    / "results"
    / "tables"
    / "Table_evidence_base_and_validation_status.xlsx"
)
PREVIEW = (
    ROOT
    / "data"
    / "exp"
    / "table_previews"
    / "Table_evidence_base_and_validation_status.png"
)

TITLE = "Evidence Base and Validation Status"

GEO = ROOT / "data" / "processed" / "geospatial"
DECISION = ROOT / "data" / "processed" / "decision"
SURVEY = ROOT / "data" / "processed" / "survey" / "vulnerability"
AUDIT = ROOT / "data" / "exp" / "data-preprocessing"
RAW_GEO = ROOT / "data" / "raw" / "geospatial"

HEADERS = [
    "Evidence layer",
    "Period",
    "Coverage",
    "Linkage",
    "Status",
    "Key limitation",
    "Use",
]

def audit_values(path: Path, key: str, value: str) -> dict[str, object]:
    frame = pd.read_csv(path)
    if key not in frame or value not in frame:
        raise RuntimeError(f"Audit schema is incomplete: {path}")
    return frame.set_index(key)[value].to_dict()


def filename_dates(paths: list[Path]) -> list[pd.Timestamp]:
    values: list[pd.Timestamp] = []
    for path in paths:
        for token in re.findall(r"20\d{2}-\d{2}-\d{2}", path.name):
            values.append(pd.Timestamp(token))
    return sorted(set(values))


def compact_period(start: pd.Timestamp, end: pd.Timestamp) -> str:
    if start.year == end.year and start.month == end.month:
        return f"{start.day}–{end.day} {start.strftime('%b %Y')}"
    return f"{start.strftime('%d %b %Y')}–{end.strftime('%d %b %Y')}"


def read_on_reference_grid(
    path: Path,
    shape: tuple[int, int],
    transform: rasterio.Affine,
    crs: rasterio.CRS,
    *,
    fill_value: float,
    resampling: Resampling,
) -> np.ndarray:
    """Read a raster on the analysis grid without assuming identical extents."""
    destination = np.full(shape, fill_value, dtype="float32")
    with rasterio.open(path) as source:
        source_values = source.read(1, masked=True).filled(fill_value).astype("float32")
        reproject(
            source=source_values,
            destination=destination,
            src_transform=source.transform,
            src_crs=source.crs,
            src_nodata=fill_value,
            dst_transform=transform,
            dst_crs=crs,
            dst_nodata=fill_value,
            resampling=resampling,
        )
    return destination


def build_rows() -> list[list[str]]:
    """Build every count and status from current source-backed artifacts."""
    road_audit = audit_values(
        AUDIT / "copernicus_road_damage_crosswalk_audit.csv",
        "measure",
        "value",
    )
    unosat_path = GEO / "reference" / "unosat_event_reference.gpkg"
    cems_path = GEO / "reference" / "copernicus_emsr927_damage_reference.gpkg"
    unosat = gpd.read_file(unosat_path, layer="affected_extent")
    cems_aois = gpd.read_file(cems_path, layer="area_of_interest")
    event_start = pd.to_datetime(unosat["event_date"]).min()

    satellite_paths = sorted((GEO / "satellite").glob("sentinel*.tif"))
    dates = filename_dates(satellite_paths)
    if len(dates) < 2:
        raise RuntimeError("Event-window satellite dates cannot be reconstructed")
    event_end = dates[-1]
    event_conditional = unosat["status"].astype(str).str.contains(
        "not_field_validated"
    ).any()

    stack_path = GEO / "hazard" / "hazard_evidence_stack_20m.tif"
    radar_path = GEO / "satellite" / "sentinel1_rtc_change_2026-08-16_2026-08-28_vh_db_20m.tif"
    optical_mask_path = GEO / "satellite" / "sentinel2_l2a_2026-08-27_valid_mask_20m.tif"
    with rasterio.open(stack_path) as dataset:
        scope = dataset.read(1, masked=True).filled(0) > 0
        stack_epsg = dataset.crs.to_epsg()
        crs_parameters = dataset.crs.to_dict()
        if stack_epsg:
            stack_crs = f"EPSG:{stack_epsg}"
        elif crs_parameters.get("proj") == "utm" and crs_parameters.get("datum") == "WGS84":
            hemisphere_base = 32700 if crs_parameters.get("south") else 32600
            stack_crs = f"EPSG:{hemisphere_base + int(crs_parameters['zone'])}"
        else:
            stack_crs = dataset.crs.to_string()
        stack_crs_object = dataset.crs
        stack_transform = dataset.transform
        stack_shape = dataset.shape
        stack_resolution = int(round(abs(dataset.transform.a)))
    radar = read_on_reference_grid(
        radar_path,
        stack_shape,
        stack_transform,
        stack_crs_object,
        fill_value=np.nan,
        resampling=Resampling.bilinear,
    )
    optical_valid = read_on_reference_grid(
        optical_mask_path,
        stack_shape,
        stack_transform,
        stack_crs_object,
        fill_value=0,
        resampling=Resampling.nearest,
    ) > 0
    with rasterio.open(optical_mask_path) as dataset:
        optical_full_grid = dataset.read(1, masked=True).filled(0) > 0
    radar_valid_pct = 100 * float(np.isfinite(radar[scope]).mean())
    optical_valid_pct = 100 * float(optical_valid[scope].mean())
    optical_full_grid_pct = 100 * float(optical_full_grid.mean())

    hydro = pd.read_csv(AUDIT / "hydrocryosphere_context_audit.csv").set_index("layer")
    river_count = int(hydro.loc["hydrorivers_context", "output_features"])
    glacier_count = int(hydro.loc["rgi7_glaciers_context", "output_features"])
    flow_paths = list((GEO / "base").glob("*flow*path*"))
    flow_status = "Flow Path ready" if flow_paths else "Flow Path pending"
    flow_limit = "D8 Flow Path available" if flow_paths else "No D8 Flow Path"

    admin_dir = RAW_GEO / "base" / "admin" / "npl_admin_boundaries_2024"
    provinces = gpd.read_file(admin_dir / "npl_admin1.shp")
    districts = gpd.read_file(admin_dir / "npl_admin2.shp")
    local_units = gpd.read_file(admin_dir / "npl_admin3.shp")
    valid_on = pd.to_datetime(districts["valid_on"]).max()
    admin_ready = all(
        frame.geometry.notna().all() for frame in (provinces, districts, local_units)
    )

    priority = pd.read_parquet(
        DECISION / "settlement_intervention_priority_preprocessed.parquet"
    )
    settlement_count = len(priority)
    baseline_reachable = int(priority["Baseline Eligible"].fillna(False).sum())
    beyond_snap = int(priority["Settlement-to-Road Snap Distance (m)"].gt(3000).sum())
    baseline_unreachable = settlement_count - baseline_reachable - beyond_snap
    osm_source = next((RAW_GEO / "osm").glob("pre_event_*/nepal-*.osm.pbf"))
    osm_dates = filename_dates([Path(osm_source.parent.name.replace("pre_event_", ""))])
    if not osm_dates:
        osm_dates = [pd.Timestamp(osm_source.parent.name.removeprefix("pre_event_"))]

    district_population = pd.read_parquet(
        GEO / "population" / "district_population_calibration.parquet"
    )
    local_population = pd.read_parquet(
        GEO / "population" / "local_unit_population_estimates.parquet"
    )
    allocation_coverage = pd.read_parquet(
        GEO
        / "population"
        / "population_allocation_threshold_summary_preprocessed.parquet"
    )
    primary_coverage = allocation_coverage.loc[
        allocation_coverage["Allocation Threshold (m)"].eq(3000)
    ]
    if len(primary_coverage) != 1:
        raise RuntimeError("Expected one primary 3 km population-coverage row")
    primary_coverage = primary_coverage.iloc[0]
    population_ready = bool(
        district_population["Calibration Difference"].abs().lt(1e-3).all()
    )
    census_year = int(local_population["Calibration Census Year"].unique().item())
    distribution_year = int(local_population["Spatial Distribution Year"].unique().item())

    nlss = pd.read_parquet(
        SURVEY / "nlss_iv_bagmati_district_vulnerability_domains_preprocessed.parquet"
    )
    nccs = pd.read_parquet(
        SURVEY / "nccs_2022_supported_domain_preparedness_preprocessed.parquet"
    )
    hrvs = pd.read_parquet(
        SURVEY / "hrvs_2016_2018_supported_domain_shock_coping_preprocessed.parquet"
    )
    no_downscaling = nccs["Settlement Downscaling"].eq("Not permitted").all() and hrvs[
        "Settlement Downscaling"
    ].eq("Not permitted").all()
    survey_status = "Context ready" if no_downscaling else "Review required"
    survey_period = (
        f"{int(hrvs['Survey Year'].min())}–{int(hrvs['Survey Year'].max())}, "
        "2022, and 2022/23"
    )

    priority_scenarios = pd.read_parquet(
        DECISION / "settlement_priority_scenarios_preprocessed.parquet",
        columns=["Scenario ID"],
    )
    allocation_priority = pd.read_parquet(
        DECISION
        / "settlement_priority_population_allocation_sensitivity_preprocessed.parquet",
        columns=["Population Allocation Specification"],
    )
    repair_candidates = pd.read_parquet(
        DECISION / "road_repair_candidate_benefits_preprocessed.parquet"
    )
    portfolio_scenarios = pd.read_parquet(
        DECISION / "road_repair_portfolio_scenarios_preprocessed.parquet"
    )
    portfolio_summary = pd.read_parquet(
        DECISION / "road_repair_portfolio_summary_preprocessed.parquet"
    )
    robustness = pd.read_parquet(
        DECISION / "settlement_priority_robustness_preprocessed.parquet"
    )
    robustness_ready = (
        robustness.loc[
            robustness["Included in Priority Ranking"].fillna(False),
            "Robustness Family Count",
        ]
        .eq(3)
        .all()
    )
    portfolio_sizes = [1, 2, 3, 5]
    scenario_group_sizes = portfolio_scenarios.groupby(
        "Repair Portfolio Size (sections)"
    ).size()
    primary_portfolios = portfolio_scenarios.loc[
        portfolio_scenarios["Primary Scenario"].fillna(False)
    ].sort_values("Repair Portfolio Size (sections)")
    portfolio_summary = portfolio_summary.sort_values(
        "Repair Portfolio Size (sections)"
    )
    portfolio_reconciles = (
        len(primary_portfolios) == len(portfolio_sizes)
        and len(portfolio_summary) == len(portfolio_sizes)
        and primary_portfolios["Repair Portfolio Size (sections)"].tolist()
        == portfolio_sizes
        and portfolio_summary["Repair Portfolio Size (sections)"].tolist()
        == portfolio_sizes
        and np.allclose(
            primary_portfolios["Portfolio Population Reconnected"],
            portfolio_summary["Primary Portfolio Population Reconnected"],
        )
        and np.allclose(
            primary_portfolios[
                "Portfolio-Weighted Finite Travel-Time Improvement (person-minutes)"
            ],
            portfolio_summary[
                "Primary Portfolio-Weighted Finite Travel-Time Improvement (person-minutes)"
            ],
        )
        and np.allclose(
            primary_portfolios["Portfolio Structural-Scenario Retention"],
            portfolio_summary["Portfolio Structural-Scenario Retention"],
        )
    )
    decision_ready = all(
        [
            priority_scenarios["Scenario ID"].nunique() == 192,
            allocation_priority["Population Allocation Specification"].nunique() == 4,
            len(allocation_priority) == settlement_count * 4,
            len(repair_candidates) == 194,
            repair_candidates["Road Repair Candidate ID"].is_unique,
            int(repair_candidates["Is Critical Road Section"].fillna(False).sum()) == 2,
            len(portfolio_scenarios) == 768,
            scenario_group_sizes.index.tolist() == portfolio_sizes,
            scenario_group_sizes.eq(192).all(),
            portfolio_reconciles,
            robustness_ready,
        ]
    )

    rows = [
        [
            "Event and damage mapping",
            compact_period(event_start, event_end),
            f"UNOSAT footprint\n{cems_aois['AOI Number'].nunique()} Copernicus AOIs",
            f"{int(road_audit['matched_cems_road_features'])}/{int(road_audit['cems_road_features'])} roads\n"
            f"{int(road_audit['matched_cems_bridge_features'])}/{int(road_audit['cems_bridge_features'])} bridges",
            "Conditional" if event_conditional else "Reference ready",
            "No field validation\nNo attribution claim" if event_conditional else "No attribution claim",
            "External event reference\nDamage-linked closures",
        ],
        [
            "Event-window satellite change",
            compact_period(dates[0], dates[-1]),
            f"Radar: {radar_valid_pct:.0f}% analysis scope\n"
            f"Optical: {optical_full_grid_pct:.1f}% grid / {optical_valid_pct:.1f}% scope",
            f"Same orbit\n{stack_resolution} m aligned grid",
            "Screening ready"
            if radar_valid_pct > 99
            and optical_full_grid_pct > 0
            and optical_valid_pct > 0
            else "Review required",
            "Threshold ≠ intensity\nCloud and snow",
            "Surface-change evidence\nMasked optical support",
        ],
        [
            "Terrain, channel, and cryosphere",
            "DEM: current\nGlaciers: ≈2000",
            f"{river_count:,} river reaches\n{glacier_count:,} glaciers",
            f"{stack_crs}\n{stack_resolution} m hazard stack",
            f"Context ready\n{flow_status}",
            f"Historical glacier extent\n{flow_limit}",
            "Slope/channel plausibility\nSource-area context",
        ],
        [
            "Administrative geography",
            f"{valid_on.year} release",
            f"{len(provinces)} provinces\n{len(districts)} districts\n{len(local_units)} local units",
            "P-code crosswalk",
            "Reporting ready" if admin_ready else "Review required",
            "No household coordinates\nNo survey downscaling",
            "Reporting domains\nSpatial aggregation",
        ],
        [
            "Road and service network",
            f"OSM: {osm_dates[0].strftime('%d %b %Y')}\nPost-event mapping",
            f"{settlement_count:,} settlements\n{baseline_reachable:,} baseline reachable",
            "Nodes and edge groups\n3 km primary snap",
            "Scenario validated" if baseline_reachable > 0 else "Review required",
            f"{baseline_unreachable:,} baseline unreachable\n{beyond_snap:,} beyond 3 km\nNo field status",
            "Closure scenarios\nService-access loss",
        ],
        [
            "Population baseline",
            f"{census_year} census totals with {distribution_year} modeled distribution",
            f"Allocated: {primary_coverage['Allocated Settlement Population']:,.0f} "
            f"({primary_coverage['Allocated Population Share (%)']:.1f}%)\n"
            f"Unallocated: {primary_coverage['Unallocated Population']:,.0f} "
            f"({primary_coverage['Unallocated Population Share (%)']:.1f}%)\n"
            f"500 m–3 km allocation: "
            f"{allocation_coverage['Allocated Population Share (%)'].min():.1f}–"
            f"{allocation_coverage['Allocated Population Share (%)'].max():.1f}%",
            f"{district_population['District'].nunique()} districts / "
            f"{local_population['Local Unit P-Code'].nunique()} local units\n≤3 km allocation",
            "Exposure ready" if population_ready else "Review required",
            "Not casualties/displacement\nUnallocated excluded",
            "Exposure and isolation\nPopulation-weighted loss",
        ],
        [
            "Household surveys",
            survey_period,
            f"{nlss['District'].nunique()} Bagmati districts\n"
            f"{nccs['Supported Domain'].nunique()} ecological-belt domains\n"
            f"{hrvs[['District', 'Survey Year']].drop_duplicates().shape[0]} historical-shock domains",
            "Supported domains only\nNo settlement linkage",
            survey_status,
            "Pre-event surveys\nMasked geocodes\nRasuwa absent from HRVS",
            "Vulnerability/preparedness\nSensitivity only",
        ],
        [
            "Decision and repair outputs",
            f"{event_end.year} event analysis",
            f"{settlement_count:,} settlements\n{priority_scenarios['Scenario ID'].nunique()} structural + "
            f"{allocation_priority['Population Allocation Specification'].nunique()} allocation specs\n"
            f"{len(repair_candidates)} candidates\n{len(portfolio_scenarios)} portfolio-scenarios",
            "Shared IDs\nSource-backed lineage checks",
            "Reproducible" if decision_ready else "Review required",
            "No cost/engineering claim\nNot a global optimum",
            "Priority robustness\nRepair screening",
        ],
    ]
    if len(rows) != 8 or any(len(row) != 7 for row in rows):
        raise RuntimeError("AnaSOP requires exactly 8 evidence rows and 7 columns")
    return rows


def build_workbook(rows: list[list[str]]) -> None:
    """Create the authoritative XLSX table with print-ready styling."""
    if len(rows) != 8 or any(len(row) != 7 for row in rows):
        raise ValueError("AnaSOP requires exactly 8 evidence rows and 7 columns")

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    PREVIEW.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Evidence audit"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A3"

    sheet.merge_cells("A1:G1")
    sheet["A1"] = TITLE
    sheet["A1"].font = Font(name="Aptos Display", size=17, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="17324D")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 32

    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=2, column=column, value=header)
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3182BD")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 34

    body_font = Font(name="Aptos", size=9, color="25313A")
    inner = Side(style="thin", color="D6DEE5")
    outer = Side(style="medium", color="535D66")
    for row_idx, values in enumerate(rows, start=3):
        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F4F7F9")
            cell.border = Border(bottom=inner)
        sheet.cell(row=row_idx, column=1).font = Font(
            name="Aptos", size=9, bold=True, color="17324D"
        )
        sheet.row_dimensions[row_idx].height = 58

    widths = [25, 21, 27, 26, 22, 29, 28]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    table = Table(displayName="EvidenceValidationAudit", ref="A2:G10")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    # Status fills are redundant to the written labels and only aid visual scanning.
    status_range = "E3:E10"
    sheet.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=['ISNUMBER(SEARCH("pending",E3))'],
            fill=PatternFill("solid", fgColor="FDE9D9"),
            font=Font(color="9C3D10", bold=True),
        ),
    )
    sheet.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=['ISNUMBER(SEARCH("conditional",E3))'],
            fill=PatternFill("solid", fgColor="FFF2CC"),
            font=Font(color="7F6000", bold=True),
        ),
    )
    sheet.conditional_formatting.add(
        status_range,
        FormulaRule(
            formula=['AND(NOT(ISNUMBER(SEARCH("pending",E3))),NOT(ISNUMBER(SEARCH("conditional",E3))))'],
            fill=PatternFill("solid", fgColor="E2F0D9"),
            font=Font(color="375623", bold=True),
        ),
    )

    # Strong frame, light internal separators.
    for col_idx in range(1, 8):
        sheet.cell(row=2, column=col_idx).border = Border(top=outer, bottom=outer)
        sheet.cell(row=10, column=col_idx).border = Border(bottom=outer)
    for row_idx in range(2, 11):
        left = sheet.cell(row=row_idx, column=1)
        right = sheet.cell(row=row_idx, column=7)
        left.border = Border(
            left=outer,
            top=left.border.top,
            bottom=left.border.bottom,
        )
        right.border = Border(
            right=outer,
            top=right.border.top,
            bottom=right.border.bottom,
        )

    sheet.auto_filter.ref = "A2:G10"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.print_area = "A1:G10"
    sheet.sheet_properties.outlinePr.summaryBelow = True
    sheet.page_margins.left = 0.2
    sheet.page_margins.right = 0.2
    sheet.page_margins.top = 0.25
    sheet.page_margins.bottom = 0.25
    sheet.page_margins.header = 0
    sheet.page_margins.footer = 0

    workbook.save(OUTPUT)


def render_preview() -> None:
    """Render the workbook's print area to a single PNG using LibreOffice."""
    soffice = shutil.which("soffice") or "/opt/homebrew/bin/soffice"
    pdftoppm = shutil.which("pdftoppm")
    if not Path(soffice).exists() or pdftoppm is None:
        raise RuntimeError("LibreOffice and pdftoppm are required for PNG rendering")

    with tempfile.TemporaryDirectory(prefix="ke03-table-preview-") as temp_name:
        temp_dir = Path(temp_name)
        profile = temp_dir / "lo-profile"
        subprocess.run(
            [
                soffice,
                f"-env:UserInstallation={profile.as_uri()}",
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                str(temp_dir),
                str(OUTPUT),
            ],
            check=True,
            capture_output=True,
            text=True,
        )
        pdf = temp_dir / f"{OUTPUT.stem}.pdf"
        if not pdf.exists():
            raise RuntimeError("LibreOffice did not create the expected PDF")
        preview_stem = temp_dir / "preview"
        subprocess.run(
            [pdftoppm, "-png", "-singlefile", "-r", "180", str(pdf), str(preview_stem)],
            check=True,
            capture_output=True,
            text=True,
        )
        rendered = Image.open(preview_stem.with_suffix(".png")).convert("RGB")
        background = Image.new("RGB", rendered.size, "white")
        content_bounds = ImageChops.difference(rendered, background).getbbox()
        if content_bounds is None:
            raise RuntimeError("Rendered PNG contains no visible table content")
        left, top, right, bottom = content_bounds
        padding = 24
        crop_box = (
            max(0, left - padding),
            max(0, top - padding),
            min(rendered.width, right + padding),
            min(rendered.height, bottom + padding),
        )
        rendered.crop(crop_box).save(PREVIEW, dpi=(180, 180))


def validate_outputs(rows: list[list[str]]) -> None:
    """Check workbook dimensions, values, formulas, and rendered output."""
    workbook = load_workbook(OUTPUT, data_only=False, read_only=False)
    sheet = workbook["Evidence audit"]
    observed_headers = [sheet.cell(row=2, column=col).value for col in range(1, 8)]
    observed_rows = [
        [sheet.cell(row=row, column=col).value for col in range(1, 8)]
        for row in range(3, 11)
    ]
    if observed_headers != HEADERS or observed_rows != rows:
        raise RuntimeError("Workbook values do not match the planned 8-by-7 table")

    errors = []
    for row in sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=7):
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("#"):
                errors.append(f"{cell.coordinate}: {value}")
            if cell.data_type == "f":
                errors.append(f"Unexpected formula in {cell.coordinate}: {value}")
    if errors:
        raise RuntimeError("Workbook validation failed: " + "; ".join(errors))
    if not PREVIEW.exists() or PREVIEW.stat().st_size < 10_000:
        raise RuntimeError("PNG preview is missing or unexpectedly small")


def main() -> None:
    rows = build_rows()
    build_workbook(rows)
    render_preview()
    validate_outputs(rows)
    print(f"Wrote {OUTPUT.relative_to(ROOT)}")
    print(f"Wrote {PREVIEW.relative_to(ROOT)}")
    print("Validated 8 rows x 7 columns; no formulas or spreadsheet error values")


if __name__ == "__main__":
    main()
