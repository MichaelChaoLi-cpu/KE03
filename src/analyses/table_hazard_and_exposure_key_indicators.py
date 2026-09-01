"""Generate the Hazard and Exposure Key Indicators table and PNG preview.

Plan: Report a compact comparison of the physical footprint, modeled
population, roads, buildings, bridges, facilities, and settlements exposed
under the three nested hazard-evidence scenarios.

Framework: AnaSOP Sections 5, 6.1, and 7 workflow steps 2-3. Hazard classes
express evidence confidence rather than physical intensity. All counts are
scenario exposure estimates, not confirmed damage, casualties, or displacement.
"""

from __future__ import annotations

from pathlib import Path
import shutil
import subprocess
import tempfile

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image, ImageChops


ROOT = Path(__file__).resolve().parents[2]
INPUT = (
    ROOT
    / "data"
    / "processed"
    / "geospatial"
    / "exposure"
    / "hazard_scenario_exposure_summary.parquet"
)
OUTPUT = (
    ROOT
    / "data"
    / "results"
    / "tables"
    / "Table_hazard_and_exposure_key_indicators.xlsx"
)
PREVIEW = (
    ROOT
    / "data"
    / "exp"
    / "table_previews"
    / "Table_hazard_and_exposure_key_indicators.png"
)

TITLE = "Hazard and Exposure Key Indicators"
HEADERS = [
    "Indicator",
    "Unit",
    "Primary\nClass 3",
    "Alternative\nClass 2",
    "Sensitivity\nClass 1",
    "Change from\nprimary",
]

SCENARIO_LABELS = {
    3: "Primary conservative",
    2: "Alternative mapped or multisensor",
    1: "Sensitivity screening",
}


def load_scenarios() -> pd.DataFrame:
    """Load and validate the three nested hazard-evidence scenarios."""
    frame = pd.read_parquet(INPUT).copy()
    required = {
        "Scenario",
        "Minimum Evidence Class",
        "Footprint Area (sq km)",
        "Exposed Population",
        "Population Share of Analysis Scope (%)",
        "Exposed Road Length (km)",
        "Buildings Intersecting Footprint",
        "Bridges Intersecting Footprint",
        "Facilities Directly Exposed",
        "Settlements Directly Exposed",
    }
    missing = required.difference(frame.columns)
    if missing:
        raise RuntimeError(f"Missing required exposure columns: {sorted(missing)}")

    frame = frame.set_index("Minimum Evidence Class").sort_index(ascending=False)
    if frame.index.tolist() != [3, 2, 1]:
        raise RuntimeError("Expected exactly the nested evidence classes 3, 2, and 1")
    for evidence_class, expected_label in SCENARIO_LABELS.items():
        if frame.loc[evidence_class, "Scenario"] != expected_label:
            raise RuntimeError(f"Unexpected label for evidence class {evidence_class}")

    monotone_columns = [
        "Footprint Area (sq km)",
        "Exposed Population",
        "Exposed Road Length (km)",
        "Buildings Intersecting Footprint",
        "Bridges Intersecting Footprint",
        "Facilities Directly Exposed",
        "Settlements Directly Exposed",
    ]
    for column in monotone_columns:
        values = frame.loc[[3, 2, 1], column].astype(float).tolist()
        if values != sorted(values):
            raise RuntimeError(f"Nested exposure is not monotone for {column}")
    return frame


def build_rows(frame: pd.DataFrame) -> list[list[object]]:
    """Construct the planned six exposure dimensions from source values."""

    def values(column: str, digits: int | None = None) -> list[float | int]:
        series = frame.loc[[3, 2, 1], column]
        if digits is None:
            return [int(round(value)) for value in series]
        return [round(float(value), digits) for value in series]

    footprint = values("Footprint Area (sq km)", 1)
    population = values("Exposed Population")
    road = values("Exposed Road Length (km)", 1)
    buildings = values("Buildings Intersecting Footprint")
    bridges = values("Bridges Intersecting Footprint")
    facilities = values("Facilities Directly Exposed")
    settlements = values("Settlements Directly Exposed")
    change = lambda series: float(series[2] / series[0] - 1)
    return [
        [
            "Hazard footprint",
            "km²",
            footprint[0],
            footprint[1],
            footprint[2],
            change(footprint),
        ],
        [
            "Exposed population",
            "people",
            population[0],
            population[1],
            population[2],
            change(population),
        ],
        [
            "Exposed road length",
            "km",
            road[0],
            road[1],
            road[2],
            change(road),
        ],
        [
            "Buildings in footprint",
            "count",
            buildings[0],
            buildings[1],
            buildings[2],
            change(buildings),
        ],
        [
            "Bridges in footprint",
            "count",
            bridges[0],
            bridges[1],
            bridges[2],
            change(bridges),
        ],
        [
            "Facilities / settlements\ndirectly exposed",
            "facility /\nsettlement count",
            f"{facilities[0]} / {settlements[0]}",
            f"{facilities[1]} / {settlements[1]}",
            f"{facilities[2]} / {settlements[2]}",
            f"+{(facilities[2] / facilities[0] - 1):.1%} / "
            f"+{(settlements[2] / settlements[0] - 1):.1%}",
        ],
    ]


def build_workbook(rows: list[list[object]]) -> None:
    """Create a compact, print-ready workbook matching the approved table style."""
    if len(rows) != 6 or any(len(row) != 6 for row in rows):
        raise RuntimeError("Revised table requires exactly 6 rows and 6 columns")

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    PREVIEW.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Hazard exposure"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A3"

    sheet.merge_cells("A1:F1")
    sheet["A1"] = TITLE
    sheet["A1"].font = Font(name="Aptos Display", size=17, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="17324D")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 32

    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=2, column=column, value=header)
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3182BD")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 40

    body_font = Font(name="Aptos", size=10, color="25313A")
    inner = Side(style="thin", color="D6DEE5")
    outer = Side(style="medium", color="535D66")
    scenario_fills = {
        3: PatternFill("solid", fgColor="E2F0D9"),
        4: PatternFill("solid", fgColor="EAF1F7"),
        5: PatternFill("solid", fgColor="FFF2CC"),
    }

    for row_idx, values in enumerate(rows, start=3):
        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if row_idx % 2 == 1:
                cell.fill = PatternFill("solid", fgColor="F4F7F9")
            cell.border = Border(bottom=inner)
        sheet.cell(row=row_idx, column=1).font = Font(
            name="Aptos", size=10, bold=True, color="17324D"
        )
        sheet.cell(row=row_idx, column=2).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        for col_idx in (3, 4, 5, 6):
            sheet.cell(row=row_idx, column=col_idx).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        for col_idx, fill in scenario_fills.items():
            sheet.cell(row=row_idx, column=col_idx).fill = fill
        sheet.row_dimensions[row_idx].height = 52

    # Store source-calculated values so non-calculating readers see complete data.
    for row_idx in range(3, 8):
        cell = sheet.cell(row=row_idx, column=6)
        cell.number_format = "0.0%"
        cell.font = Font(name="Aptos", size=10, bold=True, color="17324D")
        cell.fill = PatternFill("solid", fgColor="E8EDF2")
    sheet.cell(row=8, column=6).fill = PatternFill("solid", fgColor="E8EDF2")
    sheet.cell(row=8, column=6).font = Font(
        name="Aptos", size=10, bold=True, color="17324D"
    )

    for row_idx in range(3, 9):
        for col_idx in (3, 4, 5):
            cell = sheet.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.0" if row_idx in (3, 5) else "#,##0"

    widths = [30, 18, 21, 22, 22, 24]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    table = Table(displayName="HazardExposureIndicators", ref="A2:F8")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    for col_idx in range(1, 7):
        sheet.cell(row=2, column=col_idx).border = Border(top=outer, bottom=outer)
        sheet.cell(row=8, column=col_idx).border = Border(bottom=outer)
    for row_idx in range(2, 9):
        left = sheet.cell(row=row_idx, column=1)
        right = sheet.cell(row=row_idx, column=6)
        left.border = Border(left=outer, top=left.border.top, bottom=left.border.bottom)
        right.border = Border(right=outer, top=right.border.top, bottom=right.border.bottom)

    sheet.auto_filter.ref = "A2:F8"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.print_area = "A1:F8"
    sheet.page_margins.left = 0.2
    sheet.page_margins.right = 0.2
    sheet.page_margins.top = 0.25
    sheet.page_margins.bottom = 0.25
    sheet.page_margins.header = 0
    sheet.page_margins.footer = 0

    workbook.save(OUTPUT)


def render_preview() -> None:
    """Render the workbook print area to a tightly cropped PNG."""
    soffice = shutil.which("soffice") or "/opt/homebrew/bin/soffice"
    pdftoppm = shutil.which("pdftoppm")
    if not Path(soffice).exists() or pdftoppm is None:
        raise RuntimeError("LibreOffice and pdftoppm are required for PNG rendering")

    with tempfile.TemporaryDirectory(prefix="ke03-table-preview-") as temp_name:
        temp_dir = Path(temp_name)
        profile = temp_dir / "lo-profile"
        subprocess.run(
            [
                soffice,
                f"-env:UserInstallation={profile.as_uri()}",
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                str(temp_dir),
                str(OUTPUT),
            ],
            check=True,
            capture_output=True,
            text=True,
        )
        pdf = temp_dir / f"{OUTPUT.stem}.pdf"
        preview_stem = temp_dir / "preview"
        subprocess.run(
            [pdftoppm, "-png", "-singlefile", "-r", "180", str(pdf), str(preview_stem)],
            check=True,
            capture_output=True,
            text=True,
        )
        rendered = Image.open(preview_stem.with_suffix(".png")).convert("RGB")
        background = Image.new("RGB", rendered.size, "white")
        content_bounds = ImageChops.difference(rendered, background).getbbox()
        if content_bounds is None:
            raise RuntimeError("Rendered PNG contains no visible table content")
        left, top, right, bottom = content_bounds
        padding = 24
        rendered.crop(
            (
                max(0, left - padding),
                max(0, top - padding),
                min(rendered.width, right + padding),
                min(rendered.height, bottom + padding),
            )
        ).save(PREVIEW, dpi=(180, 180))


def validate_outputs(rows: list[list[object]]) -> None:
    """Validate source-backed values, formulas, dimensions, and preview output."""
    workbook = load_workbook(OUTPUT, data_only=False, read_only=False)
    sheet = workbook["Hazard exposure"]
    observed_headers = [sheet.cell(row=2, column=col).value for col in range(1, 7)]
    if observed_headers != HEADERS:
        raise RuntimeError("Workbook headers do not match the planned table")

    for row_idx, expected in enumerate(rows, start=3):
        for col_idx, expected_value in enumerate(expected, start=1):
            observed = sheet.cell(row=row_idx, column=col_idx).value
            if isinstance(expected_value, float):
                if not isinstance(observed, (int, float)) or abs(observed - expected_value) > 1e-12:
                    raise RuntimeError(
                        f"Unexpected value in {get_column_letter(col_idx)}{row_idx}: {observed}"
                    )
            elif observed != expected_value:
                raise RuntimeError(
                    f"Unexpected value in {get_column_letter(col_idx)}{row_idx}: {observed}"
                )

    formula_errors = []
    for row in sheet.iter_rows(min_row=1, max_row=8, min_col=1, max_col=6):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("#"):
                formula_errors.append(f"{cell.coordinate}: {cell.value}")
    if formula_errors:
        raise RuntimeError("Spreadsheet error values: " + "; ".join(formula_errors))
    data_only = load_workbook(OUTPUT, data_only=True, read_only=False)["Hazard exposure"]
    if any(data_only.cell(row=row_idx, column=6).value is None for row_idx in range(3, 8)):
        raise RuntimeError("Change-from-primary values are blank for data-only readers")
    if not PREVIEW.exists() or PREVIEW.stat().st_size < 10_000:
        raise RuntimeError("PNG preview is missing or unexpectedly small")


def main() -> None:
    scenarios = load_scenarios()
    rows = build_rows(scenarios)
    build_workbook(rows)
    render_preview()
    validate_outputs(rows)
    print(f"Wrote {OUTPUT.relative_to(ROOT)}")
    print(f"Wrote {PREVIEW.relative_to(ROOT)}")
    print("Validated 6 rows x 6 columns; machine-readable percentage changes; nested exposure monotonic")


if __name__ == "__main__":
    main()
