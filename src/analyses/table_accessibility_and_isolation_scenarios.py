"""Generate Appendix Table A3: structural scenario factor activity."""

from __future__ import annotations

from pathlib import Path
import shutil
import subprocess
import tempfile

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image, ImageChops

ROOT = Path(__file__).resolve().parents[2]
INPUT = ROOT / "data/results/supplementary/structural_scenario_factor_activity.csv"
OUTPUT = ROOT / "data/results/tables/Table_accessibility_and_isolation_scenarios.xlsx"
PREVIEW = ROOT / "data/exp/table_previews/Table_accessibility_and_isolation_scenarios.png"
TITLE = "Structural Scenario Factor Activity"
SOURCE_COLUMNS = [
    "Factor", "Levels", "Paired Comparison Groups", "Graph-Active Groups",
    "Access-Outcome-Active Groups", "Maximum Newly Isolated Population Change",
    "Maximum Delayed Population Change", "Maximum Person-Minutes Change",
    "Top-10-Set Change Groups", "Top-10-Order Change Groups",
]
HEADERS = [
    "Factor", "Levels", "Paired\ngroups", "Graph-active\ngroups", "Access-active\ngroups",
    "Maximum newly\nisolated change", "Maximum delayed\npopulation change",
    "Maximum person-minute\nchange", "Top-10 set\nchanges", "Top-10 order\nchanges",
]


def build_rows() -> list[list[object]]:
    frame = pd.read_csv(INPUT)
    if frame["Factor"].tolist() != [
        "Hazard evidence threshold", "Road closure rule", "Facility availability",
        "Topology repair threshold", "Settlement snap distance",
    ]:
        raise RuntimeError("Unexpected structural factor inventory")
    rows: list[list[object]] = []
    for _, row in frame[SOURCE_COLUMNS].iterrows():
        rows.append([
            str(row["Factor"]), str(row["Levels"]), int(row["Paired Comparison Groups"]),
            int(row["Graph-Active Groups"]), int(row["Access-Outcome-Active Groups"]),
            int(round(float(row["Maximum Newly Isolated Population Change"]))),
            int(round(float(row["Maximum Delayed Population Change"]))),
            int(round(float(row["Maximum Person-Minutes Change"]))),
            int(row["Top-10-Set Change Groups"]), int(row["Top-10-Order Change Groups"]),
        ])
    return rows


def build_workbook(rows: list[list[object]]) -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    PREVIEW.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Factor Activity"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A3"
    sheet.merge_cells("A1:J1")
    sheet["A1"] = TITLE
    sheet["A1"].font = Font(name="Aptos Display", size=17, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="17324D")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 32
    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=2, column=column, value=header)
        cell.font = Font(name="Aptos", size=9.5, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3182BD")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 48
    inner = Side(style="thin", color="D6DEE5")
    outer = Side(style="medium", color="535D66")
    for row_idx, values in enumerate(rows, start=3):
        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Aptos", size=9.2, color="25313A")
            cell.alignment = Alignment(horizontal="left" if col_idx <= 2 else "center", vertical="center", wrap_text=True)
            if row_idx % 2 == 1:
                cell.fill = PatternFill("solid", fgColor="F4F7F9")
            if col_idx == 5 and int(value) == 0:
                cell.fill = PatternFill("solid", fgColor="FCE4D6")
            cell.border = Border(bottom=inner)
        sheet.cell(row=row_idx, column=1).font = Font(name="Aptos", size=9.2, bold=True, color="17324D")
        for col_idx in range(3, 11):
            sheet.cell(row=row_idx, column=col_idx).number_format = "#,##0"
        sheet.row_dimensions[row_idx].height = 42
    widths = [27, 37, 14, 16, 16, 22, 23, 24, 15, 17]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    table = Table(displayName="StructuralFactorActivity", ref="A2:J7")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    sheet.add_table(table)
    for col_idx in range(1, 11):
        sheet.cell(row=2, column=col_idx).border = Border(top=outer, bottom=outer)
        sheet.cell(row=7, column=col_idx).border = Border(bottom=outer)
    for row_idx in range(2, 8):
        left, right = sheet.cell(row=row_idx, column=1), sheet.cell(row=row_idx, column=10)
        left.border = Border(left=outer, top=left.border.top, bottom=left.border.bottom)
        right.border = Border(right=outer, top=right.border.top, bottom=right.border.bottom)
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.print_area = "A1:J7"
    sheet.page_margins.left = sheet.page_margins.right = 0.2
    sheet.page_margins.top = sheet.page_margins.bottom = 0.25

    detail = workbook.create_sheet("Full Diagnostics")
    full = pd.read_csv(INPUT)
    for col_idx, header in enumerate(full.columns, start=1):
        detail.cell(row=1, column=col_idx, value=header)
    for row_idx, values in enumerate(full.itertuples(index=False, name=None), start=2):
        for col_idx, value in enumerate(values, start=1):
            detail.cell(row=row_idx, column=col_idx, value=value)
    detail.sheet_state = "hidden"
    workbook.save(OUTPUT)


def render_preview() -> None:
    soffice = shutil.which("soffice") or "/opt/homebrew/bin/soffice"
    pdftoppm = shutil.which("pdftoppm")
    if not Path(soffice).exists() or pdftoppm is None:
        raise RuntimeError("LibreOffice and pdftoppm are required")
    with tempfile.TemporaryDirectory(prefix="ke03-factor-table-") as temp_name:
        temp_dir = Path(temp_name)
        subprocess.run([soffice, f"-env:UserInstallation={(temp_dir / 'lo-profile').as_uri()}", "--headless", "--convert-to", "pdf", "--outdir", str(temp_dir), str(OUTPUT)], check=True, capture_output=True, text=True)
        pdf = temp_dir / f"{OUTPUT.stem}.pdf"
        stem = temp_dir / "preview"
        subprocess.run([pdftoppm, "-png", "-singlefile", "-r", "180", str(pdf), str(stem)], check=True, capture_output=True, text=True)
        rendered = Image.open(stem.with_suffix(".png")).convert("RGB")
        bounds = ImageChops.difference(rendered, Image.new("RGB", rendered.size, "white")).getbbox()
        if bounds is None:
            raise RuntimeError("Rendered preview is blank")
        left, top, right, bottom = bounds
        rendered.crop((max(0, left - 24), max(0, top - 24), min(rendered.width, right + 24), min(rendered.height, bottom + 24))).save(PREVIEW, dpi=(180, 180))


def validate_outputs(rows: list[list[object]]) -> None:
    sheet = load_workbook(OUTPUT, data_only=False)["Factor Activity"]
    observed = [[sheet.cell(row=row, column=col).value for col in range(1, 11)] for row in range(3, 8)]
    if observed != rows:
        raise RuntimeError("Workbook values changed during serialization")
    if not PREVIEW.exists() or PREVIEW.stat().st_size < 10_000:
        raise RuntimeError("PNG preview is missing or unexpectedly small")


def main() -> None:
    rows = build_rows()
    build_workbook(rows)
    render_preview()
    validate_outputs(rows)
    print(f"Wrote {OUTPUT.relative_to(ROOT)}")
    print(f"Wrote {PREVIEW.relative_to(ROOT)}")
    print("Validated five paired factor-activity rows; full diagnostics retained in hidden sheet")


if __name__ == "__main__":
    main()
