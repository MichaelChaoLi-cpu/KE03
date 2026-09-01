"""Generate the Top 10 Priority Settlements table and PNG preview.

Plan: provide a compact main-text decision summary for the ten highest-ranked
settlements, separating primary priority from scenario eligibility, structural,
population-allocation, and weighting stability, and the district-vulnerability
sensitivity ranking.

Framework: AnaSOP Sections 6.2-6.5 and workflow steps 7-9. The primary score is
the equal-weight mean of hazard, population-exposure, and accessibility
components. Vulnerability context is sensitivity-only and is not interpreted as
a settlement-level measurement.
"""

from __future__ import annotations

from pathlib import Path
import shutil
import subprocess
import tempfile

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image, ImageChops


ROOT = Path(__file__).resolve().parents[2]
PRIMARY = (
    ROOT
    / "data"
    / "processed"
    / "decision"
    / "settlement_intervention_priority_preprocessed.parquet"
)
ROBUSTNESS = (
    ROOT
    / "data"
    / "processed"
    / "decision"
    / "settlement_priority_robustness_preprocessed.parquet"
)
OUTPUT = (
    ROOT
    / "data"
    / "results"
    / "tables"
    / "Table_top_10_priority_settlements.xlsx"
)
PREVIEW = (
    ROOT
    / "data"
    / "exp"
    / "table_previews"
    / "Table_top_10_priority_settlements.png"
)

TITLE = "Top 10 Priority Settlements"
HEADERS = [
    "Settlement",
    "Hazard evidence\nclass (within 500 m)",
    "Estimated\npopulation",
    "Accessibility\nstatus",
    "Intervention\npriority",
    "Priority\nrank",
    "Scenario inclusion\nfrequency",
    "Top-10 stability\nstructural / allocation / weight / balanced",
    "Vulnerability\nsensitivity rank",
]


def build_top_ten() -> pd.DataFrame:
    """Join the registered primary ranking to its robustness diagnostics."""
    primary = pd.read_parquet(PRIMARY)
    robustness = pd.read_parquet(ROBUSTNESS)
    primary = primary.loc[
        primary["Primary Scenario"].fillna(False)
        & primary["Included in Priority Ranking"].fillna(False)
        & primary["Priority Rank"].le(10)
    ].copy()
    primary = primary.sort_values(["Priority Rank", "OSM Settlement ID"])
    if len(primary) != 10 or primary["Priority Rank"].tolist() != list(range(1, 11)):
        raise RuntimeError("Primary result must contain exactly ranks 1 through 10")

    components = primary[
        [
            "Hazard Priority Component",
            "Exposure Priority Component",
            "Accessibility Priority Component",
        ]
    ].mean(axis=1)
    if not np.allclose(components, primary["Intervention Priority"]):
        raise RuntimeError("Primary priority scores do not equal the three-component mean")

    robustness_columns = [
        "OSM Settlement ID",
        "Scenario Inclusion Frequency",
        "Structural-Scenario Top-10 Frequency",
        "Allocation-Threshold Top-10 Frequency",
        "Weight-Rule Top-10 Frequency",
        "Rank Stability",
        "Robustness Family Count",
        "Robustness Specification Count",
    ]
    top = primary.merge(
        robustness[robustness_columns],
        on="OSM Settlement ID",
        how="left",
        validate="one_to_one",
    )
    frequency_columns = [
        "Scenario Inclusion Frequency",
        "Structural-Scenario Top-10 Frequency",
        "Allocation-Threshold Top-10 Frequency",
        "Weight-Rule Top-10 Frequency",
        "Rank Stability",
    ]
    if top[frequency_columns].isna().any().any():
        raise RuntimeError("Top-ten settlements are missing robustness diagnostics")
    for column in frequency_columns:
        if not top[column].between(0, 1).all():
            raise RuntimeError(f"{column} values fall outside [0, 1]")
    expected_balanced = (
        top["Structural-Scenario Top-10 Frequency"]
        + top["Allocation-Threshold Top-10 Frequency"]
        + top["Weight-Rule Top-10 Frequency"]
    ) / 3
    if not np.allclose(expected_balanced, top["Rank Stability"]):
        raise RuntimeError("Rank Stability is not the equal-family mean")
    if not top["Robustness Family Count"].eq(3).all():
        raise RuntimeError("Top-ten settlements must have three robustness families")
    if not top["Robustness Specification Count"].eq(10203).all():
        raise RuntimeError("Unexpected robustness specification denominator")
    return top


def settlement_label(row: pd.Series) -> str:
    name = str(row["Settlement Name (English Preferred)"])
    if name.startswith("OSM Settlement "):
        identifier = name.removeprefix("OSM Settlement ")
        name = f"Unnamed settlement (OSM {identifier})"
    return f"{name}\n{row['Local Unit']}, {row['District']}"


def accessibility_label(row: pd.Series) -> str:
    status = str(row["Accessibility Status"])
    if status == "newly isolated":
        return "Newly isolated"
    if status == "delay over 5 minutes":
        return "Delayed >5 min"
    raise RuntimeError(f"Unexpected top-ten accessibility status: {status}")


def build_rows(top: pd.DataFrame) -> list[list[object]]:
    rows: list[list[object]] = []
    for _, row in top.iterrows():
        rows.append(
            [
                settlement_label(row),
                int(row["Maximum Evidence Class within 500 m"]),
                int(round(float(row["Estimated Settlement Population"]))),
                accessibility_label(row),
                float(row["Intervention Priority"]),
                int(row["Priority Rank"]),
                float(row["Scenario Inclusion Frequency"]),
                f"{row['Structural-Scenario Top-10 Frequency']:.0%} / "
                f"{row['Allocation-Threshold Top-10 Frequency']:.0%} / "
                f"{row['Weight-Rule Top-10 Frequency']:.0%} / "
                f"{row['Rank Stability']:.0%}",
                int(row["Sensitivity Priority Rank"]),
            ]
        )
    if len(rows) != 10 or any(len(row) != 9 for row in rows):
        raise RuntimeError("AnaSOP requires exactly 10 rows and 9 columns")
    return rows


def build_workbook(rows: list[list[object]]) -> None:
    """Create the authoritative XLSX using the approved compact table style."""
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    PREVIEW.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Priority Settlements"
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A3"

    sheet.merge_cells("A1:I1")
    sheet["A1"] = TITLE
    sheet["A1"].font = Font(name="Aptos Display", size=17, bold=True, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="17324D")
    sheet["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 32

    for column, header in enumerate(HEADERS, start=1):
        cell = sheet.cell(row=2, column=column, value=header)
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3182BD")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[2].height = 45

    body_font = Font(name="Aptos", size=9.5, color="25313A")
    inner = Side(style="thin", color="D6DEE5")
    outer = Side(style="medium", color="535D66")
    for row_idx, values in enumerate(rows, start=3):
        for col_idx, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if row_idx % 2 == 1:
                cell.fill = PatternFill("solid", fgColor="F4F7F9")
            cell.border = Border(bottom=inner)
        sheet.cell(row=row_idx, column=1).font = Font(
            name="Aptos", size=9.5, bold=True, color="17324D"
        )
        for col_idx in (2, 3, 4, 5, 6, 7, 8, 9):
            sheet.cell(row=row_idx, column=col_idx).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        sheet.cell(row=row_idx, column=3).number_format = "#,##0"
        sheet.cell(row=row_idx, column=5).number_format = "0.000"
        sheet.cell(row=row_idx, column=6).number_format = "0"
        sheet.cell(row=row_idx, column=7).number_format = "0%"
        sheet.cell(row=row_idx, column=9).number_format = "0"

        rank = int(values[5])
        if rank <= 3:
            sheet.cell(row=row_idx, column=6).fill = PatternFill("solid", fgColor="D9EAD3")
            sheet.cell(row=row_idx, column=6).font = Font(
                name="Aptos", size=9.5, bold=True, color="375623"
            )
        status_fill = "FCE4D6" if values[3] == "Newly isolated" else "DDEBF7"
        sheet.cell(row=row_idx, column=4).fill = PatternFill("solid", fgColor=status_fill)
        stability = float(str(values[7]).split("/")[-1].strip().rstrip("%")) / 100
        stability_fill = (
            "D9EAD3" if stability >= 0.9 else "FFF2CC" if stability >= 0.5 else "FCE4D6"
        )
        sheet.cell(row=row_idx, column=8).fill = PatternFill(
            "solid", fgColor=stability_fill
        )
        sheet.row_dimensions[row_idx].height = 42

    widths = [36, 21, 19, 20, 17, 13, 22, 34, 20]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    table = Table(displayName="TopPrioritySettlements", ref="A2:I12")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    sheet.add_table(table)

    for col_idx in range(1, 10):
        sheet.cell(row=2, column=col_idx).border = Border(top=outer, bottom=outer)
        sheet.cell(row=12, column=col_idx).border = Border(bottom=outer)
    for row_idx in range(2, 13):
        left = sheet.cell(row=row_idx, column=1)
        right = sheet.cell(row=row_idx, column=9)
        left.border = Border(left=outer, top=left.border.top, bottom=left.border.bottom)
        right.border = Border(right=outer, top=right.border.top, bottom=right.border.bottom)

    sheet.auto_filter.ref = "A2:I12"
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A3
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 1
    sheet.print_area = "A1:I12"
    sheet.page_margins.left = 0.2
    sheet.page_margins.right = 0.2
    sheet.page_margins.top = 0.25
    sheet.page_margins.bottom = 0.25
    sheet.page_margins.header = 0
    sheet.page_margins.footer = 0
    workbook.save(OUTPUT)


def render_preview() -> None:
    """Render the workbook print area to a tightly cropped PNG."""
    soffice = shutil.which("soffice") or "/opt/homebrew/bin/soffice"
    pdftoppm = shutil.which("pdftoppm")
    if not Path(soffice).exists() or pdftoppm is None:
        raise RuntimeError("LibreOffice and pdftoppm are required for PNG rendering")

    with tempfile.TemporaryDirectory(prefix="ke03-table-preview-") as temp_name:
        temp_dir = Path(temp_name)
        profile = temp_dir / "lo-profile"
        subprocess.run(
            [
                soffice,
                f"-env:UserInstallation={profile.as_uri()}",
                "--headless",
                "--convert-to",
                "pdf",
                "--outdir",
                str(temp_dir),
                str(OUTPUT),
            ],
            check=True,
            capture_output=True,
            text=True,
        )
        pdf = temp_dir / f"{OUTPUT.stem}.pdf"
        preview_stem = temp_dir / "preview"
        subprocess.run(
            ["pdftoppm", "-png", "-singlefile", "-r", "180", str(pdf), str(preview_stem)],
            check=True,
            capture_output=True,
            text=True,
        )
        rendered = Image.open(preview_stem.with_suffix(".png")).convert("RGB")
        background = Image.new("RGB", rendered.size, "white")
        content_bounds = ImageChops.difference(rendered, background).getbbox()
        if content_bounds is None:
            raise RuntimeError("Rendered PNG contains no visible table content")
        left, top, right, bottom = content_bounds
        padding = 24
        rendered.crop(
            (
                max(0, left - padding),
                max(0, top - padding),
                min(rendered.width, right + padding),
                min(rendered.height, bottom + padding),
            )
        ).save(PREVIEW, dpi=(180, 180))


def validate_outputs(rows: list[list[object]]) -> None:
    """Validate workbook values, dimensions, errors, and PNG output."""
    workbook = load_workbook(OUTPUT, data_only=False, read_only=False)
    sheet = workbook["Priority Settlements"]
    observed_headers = [sheet.cell(row=2, column=col).value for col in range(1, 10)]
    observed_rows = [
        [sheet.cell(row=row, column=col).value for col in range(1, 10)]
        for row in range(3, 13)
    ]
    if observed_headers != HEADERS:
        raise RuntimeError("Workbook headers do not match the planned table")
    for expected, observed in zip(rows, observed_rows, strict=True):
        for expected_value, observed_value in zip(expected, observed, strict=True):
            if isinstance(expected_value, float):
                if not np.isclose(expected_value, observed_value):
                    raise RuntimeError("Workbook numeric values changed during serialization")
            elif expected_value != observed_value:
                raise RuntimeError("Workbook values do not match the planned table")
    for row in sheet.iter_rows(min_row=1, max_row=12, min_col=1, max_col=9):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("#"):
                raise RuntimeError(f"Spreadsheet error value in {cell.coordinate}: {cell.value}")
            if cell.data_type == "f":
                raise RuntimeError(f"Unexpected formula in {cell.coordinate}")
    if not PREVIEW.exists() or PREVIEW.stat().st_size < 10_000:
        raise RuntimeError("PNG preview is missing or unexpectedly small")


def main() -> None:
    top = build_top_ten()
    rows = build_rows(top)
    build_workbook(rows)
    render_preview()
    validate_outputs(rows)
    print(f"Wrote {OUTPUT.relative_to(ROOT)}")
    print(f"Wrote {PREVIEW.relative_to(ROOT)}")
    print("Validated 10 rows x 9 columns; structural, allocation, weight-rule, and family-balanced stability shown")


if __name__ == "__main__":
    main()
